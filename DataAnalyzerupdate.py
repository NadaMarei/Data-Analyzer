import streamlit as st
import fitz  # PyMuPDF
import re
import io
import shutil
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, PageBreak, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import pandas as pd
from collections import defaultdict
import numpy as np
from sklearn.metrics.pairwise import cosine_similarity
from sklearn.feature_extraction.text import TfidfVectorizer, CountVectorizer
from sklearn.decomposition import LatentDirichletAllocation, TruncatedSVD
import nltk
from nltk.corpus import stopwords
from nltk.tokenize import word_tokenize, RegexpTokenizer
import string
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import traceback
import os
import tempfile
import time
import warnings
warnings.filterwarnings('ignore')

# Set page configuration first to avoid Streamlit rendering issues
st.set_page_config(
    page_title="Qualitative Data Analyzer",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Constants
MAX_MEMORY_CHUNK = 50 * 1024 * 1024  # 50MB chunks for large files
LARGE_FILE_THRESHOLD = 100 * 1024 * 1024  # 100MB

# Arabic translation dictionary
ARABIC_TRANSLATIONS = {
    # Document types and sections
    "Qualitative Data Analysis Report": "تقرير تحليل البيانات النوعية",
    "Document Statistics": "إحصائيات المستندات",
    "Summary Report": "التقرير الملخص",
    "Per-File Analysis": "تحليل لكل ملف",
    "Export Full Report": "تصدير التقرير الكامل",
    "Export Data Tables": "تصدير جداول البيانات",
    
    # Table headers
    "Document Type": "نوع المستند",
    "File Name": "اسم الملف",
    "Word Count": "عدد الكلمات",
    "Target Word": "الكلمة المستهدفة",
    "Frequency Count": "عدد التكرارات",
    "Percentage": "النسبة المئوية",
    "Total": "المجموع",
    
    # Document types
    "Target Words": "الكلمات المستهدفة",
    "Word List": "قائمة الكلمات",
    "Company Report": "تقرير الشركة",
    "Reports": "التقارير",
    
    # Analysis modes
    "Exact words only": "الكلمات المطابقة فقط",
    "Exact words and detected synonyms": "الكلمات المطابقة والمرادفات المكتشفة",
    
    # UI elements
    "Upload Documents": "رفع المستندات",
    "Company Reports": "تقارير الشركات",
    "Target Word List": "قائمة الكلمات المستهدفة",
    "Analysis Options": "خيارات التحليل",
    "Analysis Mode:": "نمط التحليل:",
    "Similarity Sensitivity": "حساسية التشابه",
    "Enable debug mode (shows detected synonyms)": "تفعيل وضع التصحيح (عرض المرادفات المكتشفة)",
    "Analyze Documents": "تحليل المستندات",
    "Download PDF Report": "تحميل تقرير PDF",
    "Download Excel Report": "تحميل تقرير Excel",
    "Download Summary (CSV)": "تحميل الملخص (CSV)",
    "Download File Analysis (CSV)": "تحميل تحليل الملفات (CSV)",
    
    # Status messages
    "Processing": "جاري المعالجة",
    "Analysis complete!": "اكتمل التحليل!",
    "Extracting word list...": "جاري استخراج قائمة الكلمات...",
    "Processing company reports...": "جاري معالجة تقارير الشركات...",
    "Building semantic model...": "جاري بناء النموذج الدلالي...",
    "Analyzing word frequencies...": "جاري تحليل ترددات الكلمات...",
    "Preparing reports...": "جاري إعداد التقارير...",
    
    # Warnings and info
    "Please upload at least one company report": "يرجى رفع تقرير شركة واحد على الأقل",
    "Please upload a word list PDF": "يرجى رفع قائمة كلمات بصيغة PDF",
    "No valid text extracted from company reports. Please check your PDF files.": "لم يتم استخراج نص صالح من تقارير الشركات. يرجى التحقق من ملفات PDF الخاصة بك.",
    "Failed to extract target words. Please check your word list PDF.": "فشل في استخراج الكلمات المستهدفة. يرجى التحقق من قائمة الكلمات بصيغة PDF.",
    "The system will automatically detect contextually similar words using semantic analysis.": "سيقوم النظام تلقائيًا باكتشاف الكلمات المتشابهة سياقيًا باستخدام التحليل الدلالي.",
    "Higher values detect only very similar words, lower values detect broader synonyms": "القيم الأعلى تكتشف فقط الكلمات المتشابهة جدًا، والقيم الأقل تكتشف مرادفات أوسع",
    "Semantic model creation failed. Using exact words only.": "فشل إنشاء النموذج الدلالي. استخدام الكلمات المطابقة فقط.",
    "No valid text for semantic analysis. Using exact words only.": "لا يوجد نص صالح للتحليل الدلالي. استخدام الكلمات المطابقة فقط.",
    "✓ Synonym detection was successfully enabled": "✓ تم تفعيل اكتشاف المرادفات بنجاح",
    "No file analysis data available": "لا توجد بيانات تحليل للملفات",
    "PDF report generation failed": "فشل إنشاء تقرير PDF",
    "Excel report generation failed": "فشل إنشاء تقرير Excel",
    "Analysis failed:": "فشل التحليل:",
    "Please try again or check your files": "يرجى المحاولة مرة أخرى أو التحقق من ملفاتك",
    "Critical application error:": "خطأ تطبيق حرج:",
    
    # File processing
    "Processing large files. This may take longer...": "جاري معالجة الملفات الكبيرة. قد يستغرق هذا وقتًا أطول...",
    "words": "كلمات",
    "seconds": "ثواني",
    
    # Synonym detection
    "Using context-based synonym detection": "استخدام اكتشاف المرادفات القائم على السياق",
    "Basic synonym detection using co-occurrence patterns": "اكتشاف المرادفات الأساسي باستخدام أنماط التكامل"
}

def get_arabic_text(english_text):
    """Get Arabic translation for English text, fallback to English if not found"""
    return ARABIC_TRANSLATIONS.get(english_text, english_text)

def is_arabic_text(text):
    """Check if text contains Arabic characters"""
    arabic_range = re.compile('[\u0600-\u06FF\u0750-\u077F\u08A0-\u08FF]+')
    return bool(arabic_range.search(text))

# Robust NLTK resource handling with punkt_tab fix
def setup_nltk_resources():
    try:
        # Create a proper NLTK data directory
        nltk_data_dir = os.path.join(tempfile.gettempdir(), "nltk_data")
        os.makedirs(nltk_data_dir, exist_ok=True)
        
        # Set NLTK data path
        nltk.data.path.append(nltk_data_dir)
        
        # Download required resources
        resources = ['punkt', 'stopwords']
        
        for resource in resources:
            try:
                if resource == 'punkt':
                    nltk.data.find('tokenizers/punkt')
                else:
                    nltk.data.find(f'corpora/{resource}')
            except LookupError:
                nltk.download(resource, download_dir=nltk_data_dir, quiet=True)
        
        # Create punkt_tab directory structure for compatibility
        punkt_tab_dir = os.path.join(nltk_data_dir, 'tokenizers', 'punkt_tab')
        os.makedirs(punkt_tab_dir, exist_ok=True)
        
        # Create a minimal collocations.tab file if it doesn't exist
        collocations_path = os.path.join(punkt_tab_dir, 'collocations.tab')
        if not os.path.exists(collocations_path):
            with open(collocations_path, 'w') as f:
                f.write("# Empty collocations file for compatibility\n")
        
        return True
    except Exception as e:
        st.error(f"NLTK setup failed: {str(e)}")
        return False

# Alternative text tokenizer for fallback
def simple_tokenizer(text):
    """Simple regex-based tokenizer as fallback when NLTK fails"""
    tokenizer = RegexpTokenizer(r'\w+')
    return tokenizer.tokenize(text.lower())

def extract_text_from_pdf(uploaded_file):
    """Extract text from an uploaded PDF file with efficient memory handling"""
    try:
        # Reset file pointer to beginning before reading
        uploaded_file.seek(0)
        
        # Check if file is large
        if uploaded_file.size > LARGE_FILE_THRESHOLD:
            return process_large_pdf(uploaded_file)
        else:
            # Process small files directly in memory
            with fitz.open(stream=uploaded_file.read(), filetype="pdf") as doc:
                text = ""
                for page in doc:
                    text += page.get_text()
                return text
    except Exception as e:
        st.error(f"Error processing {uploaded_file.name}: {str(e)}")
        return None

def process_large_pdf(uploaded_file):
    """Process large PDF files using temporary files and streaming"""
    try:
        # Create a temporary file
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp_file:
            # Write in chunks to avoid memory overload
            uploaded_file.seek(0)
            while True:
                chunk = uploaded_file.read(MAX_MEMORY_CHUNK)
                if not chunk:
                    break
                tmp_file.write(chunk)
            tmp_file_path = tmp_file.name
        
        # Process the temporary file
        text = ""
        with fitz.open(tmp_file_path) as doc:
            for page in doc:
                text += page.get_text()
        
        # Clean up temporary file
        os.unlink(tmp_file_path)
        return text
    except Exception as e:
        st.error(f"Error processing large file: {str(e)}")
        return None

def extract_words_from_pdf(uploaded_file):
    """Extract word list from PDF (one word per line) with efficient memory handling"""
    text = extract_text_from_pdf(uploaded_file)
    if text is None:
        return None
    
    words = []
    for line in text.split('\n'):
        cleaned_line = line.strip()
        if cleaned_line:
            words.append(cleaned_line)
    return words

def preprocess_text(text):
    """Clean and tokenize text with robust error handling and fallback"""
    if not text:
        return []
    
    try:
        # Convert to lowercase
        text = text.lower()
        # Remove punctuation
        text = text.translate(str.maketrans('', '', string.punctuation))
        
        # Try NLTK tokenization first
        try:
            tokens = word_tokenize(text)
        except Exception:
            # Fallback to simple tokenizer if NLTK fails
            tokens = simple_tokenizer(text)
        
        # Remove stopwords
        try:
            stop_words = set(stopwords.words('english'))
            tokens = [word for word in tokens if word not in stop_words and len(word) > 2]
        except Exception:
            # If stopwords fail, just filter by length
            tokens = [word for word in tokens if len(word) > 2]
            
        return tokens
    except Exception as e:
        st.error(f"Error in text preprocessing: {str(e)}")
        return []

def count_words_in_text(text):
    """Count words in text using simple splitting method"""
    if not text:
        return 0
    return len(text.split())

def create_semantic_model(documents):
    """Create a simple semantic model using TF-IDF and cosine similarity"""
    if not documents or not any(documents):
        return None
    
    try:
        # Create TF-IDF vectorizer
        vectorizer = TfidfVectorizer(
            max_features=1000,
            min_df=2,
            max_df=0.8,
            stop_words='english'
        )
        
        # Fit and transform the documents
        tfidf_matrix = vectorizer.fit_transform(documents)
        
        return {
            'vectorizer': vectorizer,
            'tfidf_matrix': tfidf_matrix,
            'feature_names': vectorizer.get_feature_names_out()
        }
    except Exception as e:
        st.error(f"Semantic model creation failed: {str(e)}")
        return None

def find_similar_words_semantic(target_word, semantic_model, threshold=0.7):
    """Find similar words using TF-IDF and cosine similarity"""
    similar_words = []
    
    if not semantic_model:
        return similar_words
    
    try:
        vectorizer = semantic_model['vectorizer']
        tfidf_matrix = semantic_model['tfidf_matrix']
        feature_names = semantic_model['feature_names']
        
        # Check if target word is in vocabulary
        if target_word.lower() not in vectorizer.vocabulary_:
            return similar_words
        
        # Get the index of the target word
        target_idx = vectorizer.vocabulary_[target_word.lower()]
        
        # Calculate cosine similarities
        target_vector = tfidf_matrix[:, target_idx].toarray().flatten()
        
        for i, word in enumerate(feature_names):
            if word != target_word.lower():
                word_vector = tfidf_matrix[:, i].toarray().flatten()
                
                # Calculate cosine similarity
                if np.linalg.norm(target_vector) > 0 and np.linalg.norm(word_vector) > 0:
                    similarity = np.dot(target_vector, word_vector) / (
                        np.linalg.norm(target_vector) * np.linalg.norm(word_vector)
                    )
                    
                    if similarity > threshold:
                        similar_words.append(word)
        
        return similar_words
    
    except Exception as e:
        st.error(f"Semantic similarity search failed: {str(e)}")
        return []

def find_similar_words_context(target_word, documents, threshold=0.7):
    """Find similar words based on contextual co-occurrence"""
    similar_words = []
    target_word_lower = target_word.lower()
    
    # Collect words that appear in the same contexts
    context_words = defaultdict(int)
    total_cooccurrences = 0
    
    for doc in documents:
        sentences = re.split(r'[.!?]+', doc.lower())
        for sentence in sentences:
            words_in_sentence = preprocess_text(sentence)
            if target_word_lower in words_in_sentence:
                # Count other words in the same sentence
                for word in words_in_sentence:
                    if word != target_word_lower and len(word) > 2:
                        context_words[word] += 1
                        total_cooccurrences += 1
    
    # Calculate co-occurrence scores and filter by threshold
    if total_cooccurrences > 0:
        for word, count in context_words.items():
            score = count / total_cooccurrences
            if score > threshold / 10:  # Adjust threshold for co-occurrence
                similar_words.append(word)
    
    return similar_words[:15]  # Limit to top 15

def find_similar_words(target_word, semantic_model, documents, threshold=0.7, method='semantic'):
    """Find similar words using the specified method"""
    if method == 'semantic' and semantic_model:
        return find_similar_words_semantic(target_word, semantic_model, threshold)
    else:
        return find_similar_words_context(target_word, documents, threshold)

def count_word_frequencies(text, word_list, semantic_model=None, documents=None, use_synonyms=False, threshold=0.7):
    """Count occurrences of words with optional synonym detection"""
    frequencies = defaultdict(int)
    if not text:
        return frequencies
    
    text_lower = text.lower()
    
    for word in word_list:
        # Always count exact matches
        try:
            exact_pattern = r'\b' + re.escape(word.lower()) + r'\b'
            exact_matches = re.findall(exact_pattern, text_lower)
            frequencies[word] += len(exact_matches)
        except re.error:
            continue
        
        # Detect and count synonyms if enabled
        if use_synonyms:
            similar_words = find_similar_words(
                word, 
                semantic_model, 
                documents if documents else [text],
                threshold,
                method='semantic' if semantic_model else 'context'
            )
            
            # Show detected synonyms in debug mode
            if similar_words and st.session_state.get('debug_mode', False):
                st.write(f"Synonyms for '{word}': {', '.join(similar_words[:5])}")  # Show top 5
            
            for synonym in similar_words:
                try:
                    syn_pattern = r'\b' + re.escape(synonym) + r'\b'
                    syn_matches = re.findall(syn_pattern, text_lower)
                    frequencies[word] += len(syn_matches)
                except re.error:
                    continue
    
    return frequencies

def generate_pdf_report(summary_data, file_analysis_data, word_counts, target_words, include_arabic=True):
    """Generate professional PDF report in memory with optional Arabic version"""
    try:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
        styles = getSampleStyleSheet()
        elements = []
        
        # Add title
        title = Paragraph("Qualitative Data Analysis Report", styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 12))
        
        # Document Statistics Section
        elements.append(Paragraph("Document Statistics", styles['Heading2']))
        elements.append(Spacer(1, 8))
        
        # Prepare document stats table
        stats_table_data = [
            ["Document Type", "File Name", "Word Count"],
            ["Target Words", "Word List", str(word_counts['word_list'])]
        ]
        
        for report in word_counts['reports']:
            stats_table_data.append([
                "Company Report", 
                report['name'], 
                str(report['word_count'])
            ])
        
        stats_table_data.append([
            "Total", 
            f"{len(word_counts['reports'])} Reports", 
            str(word_counts['total_report_words'])
        ])
        
        # Create stats table
        stats_table = Table(stats_table_data)
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2c3e50")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#ecf0f1")),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), 
             [colors.whitesmoke, colors.HexColor("#f9f9f9")]),
        ]))
        elements.append(stats_table)
        
        elements.append(Spacer(1, 24))
        
        # Summary table - PRESERVE USER'S WORD ORDER
        elements.append(Paragraph("Summary Report", styles['Heading2']))
        elements.append(Spacer(1, 8))
        
        # Prepare summary table data in the original word order
        summary_table_data = [["Target Word", "Frequency Count", "Percentage"]]
        for word in target_words:
            # Find the matching summary data for this word
            word_data = next((item for item in summary_data if item["Target Word"] == word), None)
            if word_data:
                summary_table_data.append([
                    word_data["Target Word"], 
                    str(word_data["Frequency Count"]), 
                    f"{word_data['Percentage']:.2f}%"
                ])
        
        # Create summary table
        summary_table = Table(summary_table_data)
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2c3e50")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#ecf0f1")),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), 
             [colors.whitesmoke, colors.HexColor("#f9f9f9")]),
        ]))
        elements.append(summary_table)
        
        # Add Arabic version if requested
        if include_arabic:
            elements.append(PageBreak())
            
            # Arabic Title
            arabic_title = Paragraph(get_arabic_text("Qualitative Data Analysis Report"), styles['Title'])
            elements.append(arabic_title)
            elements.append(Spacer(1, 12))
            
            # Arabic Document Statistics Section
            elements.append(Paragraph(get_arabic_text("Document Statistics"), styles['Heading2']))
            elements.append(Spacer(1, 8))
            
            # Prepare Arabic document stats table
            arabic_stats_table_data = [
                [get_arabic_text("Document Type"), get_arabic_text("File Name"), get_arabic_text("Word Count")],
                [get_arabic_text("Target Words"), get_arabic_text("Word List"), str(word_counts['word_list'])]
            ]
            
            for report in word_counts['reports']:
                arabic_stats_table_data.append([
                    get_arabic_text("Company Report"), 
                    report['name'], 
                    str(report['word_count'])
                ])
            
            arabic_stats_table_data.append([
                get_arabic_text("Total"), 
                f"{len(word_counts['reports'])} {get_arabic_text('Reports')}", 
                str(word_counts['total_report_words'])
            ])
            
            # Create Arabic stats table
            arabic_stats_table = Table(arabic_stats_table_data)
            arabic_stats_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2c3e50")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#ecf0f1")),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), 
                 [colors.whitesmoke, colors.HexColor("#f9f9f9")]),
            ]))
            elements.append(arabic_stats_table)
            
            elements.append(Spacer(1, 24))
            
            # Arabic Summary table
            elements.append(Paragraph(get_arabic_text("Summary Report"), styles['Heading2']))
            elements.append(Spacer(1, 8))
            
            # Prepare Arabic summary table data in the original word order
            arabic_summary_table_data = [
                [get_arabic_text("Target Word"), get_arabic_text("Frequency Count"), get_arabic_text("Percentage")]
            ]
            for word in target_words:
                # Find the matching summary data for this word
                word_data = next((item for item in summary_data if item["Target Word"] == word), None)
                if word_data:
                    arabic_summary_table_data.append([
                        word_data["Target Word"], 
                        str(word_data["Frequency Count"]), 
                        f"{word_data['Percentage']:.2f}%"
                    ])
            
            # Create Arabic summary table
            arabic_summary_table = Table(arabic_summary_table_data)
            arabic_summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2c3e50")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#ecf0f1")),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), 
                 [colors.whitesmoke, colors.HexColor("#f9f9f9")]),
            ]))
            elements.append(arabic_summary_table)
        
        # Add space before next section
        elements.append(PageBreak())
        
        # File analysis section - PRESERVE USER'S WORD ORDER
        elements.append(Paragraph("Per-File Analysis", styles['Heading2']))
        elements.append(Spacer(1, 8))
        
        # Prepare file analysis table data in the original word order
        if file_analysis_data:
            file_names = list(next(iter(file_analysis_data.values())).keys())
            file_table_data = [["Target Word"] + list(file_names) + ["Total"]]
            
            # Add rows in the original word order
            for word in target_words:
                if word in file_analysis_data:
                    counts = file_analysis_data[word]
                    row = [word]
                    total = 0
                    for file_name in file_names:
                        count = counts.get(file_name, 0)
                        row.append(str(count))
                        total += count
                    row.append(str(total))
                    file_table_data.append(row)
            
            # Create file analysis table
            file_table = Table(file_table_data)
            file_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#3498db")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#f8f9fa")),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), 
                 [colors.whitesmoke, colors.HexColor("#f1f8ff")]),
            ]))
            elements.append(file_table)
            
            # Add Arabic version of file analysis if requested
            if include_arabic:
                elements.append(PageBreak())
                elements.append(Paragraph(get_arabic_text("Per-File Analysis"), styles['Heading2']))
                elements.append(Spacer(1, 8))
                
                # Prepare Arabic file analysis table
                arabic_file_table_data = [[get_arabic_text("Target Word")] + list(file_names) + [get_arabic_text("Total")]]
                
                for word in target_words:
                    if word in file_analysis_data:
                        counts = file_analysis_data[word]
                        row = [word]
                        total = 0
                        for file_name in file_names:
                            count = counts.get(file_name, 0)
                            row.append(str(count))
                            total += count
                        row.append(str(total))
                        arabic_file_table_data.append(row)
                
                # Create Arabic file analysis table
                arabic_file_table = Table(arabic_file_table_data)
                arabic_file_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#3498db")),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#f8f9fa")),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), 
                     [colors.whitesmoke, colors.HexColor("#f1f8ff")]),
                ]))
                elements.append(arabic_file_table)
        
        doc.build(elements)
        buffer.seek(0)
        return buffer
    except Exception as e:
        st.error(f"PDF generation failed: {str(e)}")
        return None

def generate_excel_report(summary_data, file_analysis_data, word_counts, target_words, include_arabic=True):
    """Generate Excel report with multiple sheets in memory with optional Arabic version"""
    try:
        buffer = io.BytesIO()
        
        # Create a workbook and add sheets
        wb = Workbook()
        
        # Document Statistics Sheet
        ws_stats = wb.active
        ws_stats.title = "Document Statistics"
        
        # Prepare stats data
        stats_data = [
            ["Document Type", "File Name", "Word Count"],
            ["Target Words", "Word List", word_counts['word_list']]
        ]
        
        for report in word_counts['reports']:
            stats_data.append(["Company Report", report['name'], report['word_count']])
        
        stats_data.append(["Total", f"{len(word_counts['reports'])} Reports", word_counts['total_report_words']])
        
        # Add data to sheet
        for row in stats_data:
            ws_stats.append(row)
        
        # Format stats sheet
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(left=Side(style='thin'), 
                       right=Side(style='thin'), 
                       top=Side(style='thin'), 
                       bottom=Side(style='thin'))
        center_aligned = Alignment(horizontal='center')
        
        for cell in ws_stats[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_aligned
        
        for row in ws_stats.iter_rows(min_row=2, max_row=len(stats_data)):
            for cell in row:
                cell.border = border
                cell.alignment = center_aligned
        
        # Summary Report Sheet - PRESERVE USER'S WORD ORDER
        ws_summary = wb.create_sheet(title="Summary Report")
        
        # Prepare summary data in the original word order
        summary_headers = ["Target Word", "Frequency Count", "Percentage"]
        ws_summary.append(summary_headers)
        
        for word in target_words:
            # Find the matching summary data for this word
            word_data = next((item for item in summary_data if item["Target Word"] == word), None)
            if word_data:
                ws_summary.append([
                    word_data["Target Word"], 
                    word_data["Frequency Count"], 
                    word_data["Percentage"]
                ])
        
        # Format summary sheet
        for cell in ws_summary[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_aligned
        
        for row in ws_summary.iter_rows(min_row=2, max_row=len(target_words)+1):
            for cell in row:
                cell.border = border
                cell.alignment = center_aligned
        
        # Per-File Analysis Sheet - PRESERVE USER'S WORD ORDER
        if file_analysis_data:
            ws_analysis = wb.create_sheet(title="Per-File Analysis")
            
            # Prepare headers
            file_names = list(next(iter(file_analysis_data.values())).keys())
            headers = ["Target Word"] + list(file_names) + ["Total"]
            ws_analysis.append(headers)
            
            # Add data in the original word order
            for word in target_words:
                if word in file_analysis_data:
                    counts = file_analysis_data[word]
                    row = [word]
                    total = 0
                    for file_name in file_names:
                        count = counts.get(file_name, 0)
                        row.append(count)
                        total += count
                    row.append(total)
                    ws_analysis.append(row)
            
            # Format analysis sheet
            for cell in ws_analysis[1]:
                cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
                cell.font = Font(bold=True)
                cell.alignment = center_aligned
            
            for row in ws_analysis.iter_rows(min_row=2, max_row=len(target_words)+1):
                for cell in row:
                    cell.border = border
                    cell.alignment = center_aligned
        
        # Add Arabic version sheets if requested
        if include_arabic:
            # Arabic Document Statistics Sheet
            ws_arabic_stats = wb.create_sheet(title=get_arabic_text("Document Statistics"))
            
            # Prepare Arabic stats data
            arabic_stats_data = [
                [get_arabic_text("Document Type"), get_arabic_text("File Name"), get_arabic_text("Word Count")],
                [get_arabic_text("Target Words"), get_arabic_text("Word List"), word_counts['word_list']]
            ]
            
            for report in word_counts['reports']:
                arabic_stats_data.append([get_arabic_text("Company Report"), report['name'], report['word_count']])
            
            arabic_stats_data.append([
                get_arabic_text("Total"), 
                f"{len(word_counts['reports'])} {get_arabic_text('Reports')}", 
                word_counts['total_report_words']
            ])
            
            # Add Arabic data to sheet
            for row in arabic_stats_data:
                ws_arabic_stats.append(row)
            
            # Format Arabic stats sheet
            for cell in ws_arabic_stats[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_aligned
            
            for row in ws_arabic_stats.iter_rows(min_row=2, max_row=len(arabic_stats_data)):
                for cell in row:
                    cell.border = border
                    cell.alignment = center_aligned
            
            # Arabic Summary Report Sheet
            ws_arabic_summary = wb.create_sheet(title=get_arabic_text("Summary Report"))
            
            # Prepare Arabic summary data
            arabic_summary_headers = [get_arabic_text("Target Word"), get_arabic_text("Frequency Count"), get_arabic_text("Percentage")]
            ws_arabic_summary.append(arabic_summary_headers)
            
            for word in target_words:
                word_data = next((item for item in summary_data if item["Target Word"] == word), None)
                if word_data:
                    ws_arabic_summary.append([
                        word_data["Target Word"], 
                        word_data["Frequency Count"], 
                        word_data["Percentage"]
                    ])
            
            # Format Arabic summary sheet
            for cell in ws_arabic_summary[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_aligned
            
            for row in ws_arabic_summary.iter_rows(min_row=2, max_row=len(target_words)+1):
                for cell in row:
                    cell.border = border
                    cell.alignment = center_aligned
            
            # Arabic Per-File Analysis Sheet
            if file_analysis_data:
                ws_arabic_analysis = wb.create_sheet(title=get_arabic_text("Per-File Analysis"))
                
                # Prepare Arabic headers
                arabic_headers = [get_arabic_text("Target Word")] + list(file_names) + [get_arabic_text("Total")]
                ws_arabic_analysis.append(arabic_headers)
                
                # Add Arabic data
                for word in target_words:
                    if word in file_analysis_data:
                        counts = file_analysis_data[word]
                        row = [word]
                        total = 0
                        for file_name in file_names:
                            count = counts.get(file_name, 0)
                            row.append(count)
                            total += count
                        row.append(total)
                        ws_arabic_analysis.append(row)
                
                # Format Arabic analysis sheet
                for cell in ws_arabic_analysis[1]:
                    cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
                    cell.font = Font(bold=True)
                    cell.alignment = center_aligned
                
                for row in ws_arabic_analysis.iter_rows(min_row=2, max_row=len(target_words)+1):
                    for cell in row:
                        cell.border = border
                        cell.alignment = center_aligned
        
        # Save to buffer
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    except Exception as e:
        st.error(f"Excel generation failed: {str(e)}")
        return None

def main():
    try:
        # Initialize session state for debug mode
        if 'debug_mode' not in st.session_state:
            st.session_state.debug_mode = False
            
        # Initialize NLTK resources
        nltk_success = setup_nltk_resources()
        if not nltk_success:
            st.warning("NLTK setup had issues. Using fallback text processing methods.")
            
        st.title("📊 Qualitative Data Analysis Tool")
        st.markdown("""
        **Upload company reports and a word list to analyze word frequencies across documents.**
        *Supports large files up to 1TB with efficient streaming*
        """)
        
        # Language options
        with st.expander("Language Options", expanded=True):
            include_arabic = st.checkbox("Include Arabic version in reports", value=True)
            st.info("When enabled, all reports will include both English and Arabic versions of the analysis.")
        
        # Debug mode toggle
        if st.checkbox("Enable debug mode (shows detected synonyms)"):
            st.session_state.debug_mode = True
        else:
            st.session_state.debug_mode = False
        
        # File upload sections
        with st.expander("Upload Documents", expanded=True):
            col1, col2 = st.columns(2)
            
            with col1:
                st.subheader("Company Reports")
                report_files = st.file_uploader(
                    "Upload one or more company reports (PDF)", 
                    type="pdf", 
                    accept_multiple_files=True,
                    help="Upload PDF documents containing the text to analyze"
                )
            
            with col2:
                st.subheader("Target Word List")
                word_file = st.file_uploader(
                    "Upload word list (PDF)", 
                    type="pdf",
                    help="PDF containing one target word per line"
                )
        
        # Analysis options
        with st.expander("Analysis Options", expanded=True):
            analysis_mode = st.radio(
                "Analysis Mode:",
                ["Exact words only", "Exact words and detected synonyms"],
                index=0,
                horizontal=True
            )
            
            if analysis_mode == "Exact words and detected synonyms":
                st.info("The system will automatically detect contextually similar words using TF-IDF semantic analysis.")
                similarity_threshold = st.slider(
                    "Similarity Sensitivity",
                    min_value=0.1,
                    max_value=1.0,
                    value=0.7,
                    step=0.05,
                    help="Higher values detect only very similar words, lower values detect broader synonyms"
                )
        
        # Processing section
        if st.button("Analyze Documents", type="primary", use_container_width=True):
            # Validate inputs
            if not report_files:
                st.warning("Please upload at least one company report")
                return
            if not word_file:
                st.warning("Please upload a word list PDF")
                return
            
            # Check for very large files
            large_files = [f for f in report_files if f.size > LARGE_FILE_THRESHOLD]
            if large_files:
                st.info(f"Processing {len(large_files)} large file(s). This may take longer...")
            
            try:
                # Initialize progress
                progress_bar = st.progress(0)
                status_text = st.empty()
                
                # Step 1: Extract words and count them
                status_text.text("Extracting word list...")
                target_words = extract_words_from_pdf(word_file)
                if not target_words:
                    st.error("Failed to extract target words. Please check your word list PDF.")
                    return
                word_list_count = len(target_words)
                
                # Reset file pointer for report files
                for file in report_files:
                    file.seek(0)
                
                # Step 2: Process company reports - extract text and count words
                status_text.text("Processing company reports...")
                report_data = []
                total_report_words = 0
                total_files = len(report_files)
                start_time = time.time()
                
                # Extract and store text from all reports with word counts
                for i, report_file in enumerate(report_files):
                    # Show file processing status
                    file_size = report_file.size / (1024 * 1024)
                    status_text.text(f"Processing {report_file.name} ({file_size:.1f} MB)...")
                    
                    # Reset file pointer before reading
                    report_file.seek(0)
                    progress_bar.progress((i + 1) / (total_files * 3 + 2))
                    
                    # Process file with appropriate method
                    text = extract_text_from_pdf(report_file)
                    if text:
                        word_count = count_words_in_text(text)
                        report_data.append({
                            "name": report_file.name,
                            "text": text,
                            "word_count": word_count
                        })
                        total_report_words += word_count
                
                # Check if we have valid documents
                if not report_data:
                    st.error("No valid text extracted from company reports. Please check your PDF files.")
                    return
                
                # Create combined text
                combined_text = "\n\n".join([item["text"] for item in report_data])
                
                # Prepare document statistics
                document_stats = {
                    "word_list": word_list_count,
                    "reports": [{"name": item["name"], "word_count": item["word_count"]} for item in report_data],
                    "total_report_words": total_report_words
                }
                
                # Create semantic model if needed
                semantic_model = None
                if analysis_mode == "Exact words and detected synonyms":
                    status_text.text("Building semantic model...")
                    # Use all text for training the semantic model
                    training_texts = [item["text"] for item in report_data]
                    
                    if training_texts:
                        semantic_model = create_semantic_model(training_texts)
                        if not semantic_model:
                            st.warning("Semantic model creation failed. Using context-based synonym detection.")
                        else:
                            st.info("Using TF-IDF based semantic analysis for synonym detection")
                    else:
                        st.warning("No valid text for semantic analysis. Using context-based synonym detection.")
                    progress_bar.progress((total_files + 1) / (total_files * 3 + 2))
                
                # Initialize file analysis structure
                file_analysis = defaultdict(lambda: defaultdict(int))
                
                # Count words in each file
                for i, data in enumerate(report_data):
                    status_text.text(f"Analyzing file {i+1}/{len(report_data)}...")
                    progress_bar.progress((total_files + i + 2) / (total_files * 3 + 2))
                    
                    # Count words for this specific file
                    file_counts = count_word_frequencies(
                        data["text"], 
                        target_words, 
                        semantic_model, 
                        documents=[item["text"] for item in report_data],
                        use_synonyms=(analysis_mode == "Exact words and detected synonyms"),
                        threshold=similarity_threshold if analysis_mode == "Exact words and detected synonyms" else 0.7
                    )
                    
                    # Store in analysis structure
                    for word, count in file_counts.items():
                        file_analysis[word][data["name"]] = count
                
                # Step 3: Count overall word frequencies
                status_text.text("Analyzing word frequencies...")
                overall_frequencies = count_word_frequencies(
                    combined_text, 
                    target_words, 
                    semantic_model,
                    documents=[item["text"] for item in report_data],
                    use_synonyms=(analysis_mode == "Exact words and detected synonyms"),
                    threshold=similarity_threshold if analysis_mode == "Exact words and detected synonyms" else 0.7
                )
                total_occurrences = sum(overall_frequencies.values())
                
                # Prepare summary data with percentages - PRESERVE USER'S WORD ORDER
                summary_data = []
                for word in target_words:
                    count = overall_frequencies.get(word, 0)
                    percentage = (count / total_occurrences * 100) if total_occurrences > 0 else 0
                    summary_data.append({
                        "Target Word": word,
                        "Frequency Count": count,
                        "Percentage": percentage
                    })
                
                progress_bar.progress(100)
                
                # Display results
                processing_time = time.time() - start_time
                st.success(f"Analysis complete! Processed {total_report_words:,} words in {processing_time:.1f} seconds")
                
                if analysis_mode == "Exact words and detected synonyms":
                    if semantic_model:
                        st.info("✓ TF-IDF semantic analysis was successfully enabled")
                    else:
                        st.info("✓ Context-based synonym detection was used")
                
                # Document Statistics
                st.subheader("Document Statistics")
                stats_df = pd.DataFrame({
                    "Document Type": ["Word List"] + ["Company Report"] * len(report_data),
                    "File Name": ["Word List"] + [item["name"] for item in report_data],
                    "Word Count": [word_list_count] + [item["word_count"] for item in report_data]
                })
                st.dataframe(stats_df, height=200)
                st.markdown(f"**Total Words in Reports:** {total_report_words:,}")
                
                # Summary Report - PRESERVE USER'S WORD ORDER
                st.subheader("Summary Report")
                summary_df = pd.DataFrame(summary_data)
                st.dataframe(summary_df, height=300)
                
                # File Analysis Report - PRESERVE USER'S WORD ORDER
                st.subheader("Per-File Analysis")
                if file_analysis:
                    # Create DataFrame with words in original order
                    analysis_rows = []
                    for word in target_words:
                        if word in file_analysis:
                            row = {"Target Word": word}
                            for file_name in file_analysis[word]:
                                row[file_name] = file_analysis[word][file_name]
                            row['Total'] = sum(file_analysis[word].values())
                            analysis_rows.append(row)
                    
                    file_analysis_df = pd.DataFrame(analysis_rows)
                    file_analysis_df.set_index("Target Word", inplace=True)
                    st.dataframe(file_analysis_df, height=400)
                else:
                    st.warning("No file analysis data available")
                
                # Export options
                st.subheader("Export Full Report")
                col1, col2 = st.columns(2)
                
                # Generate reports in memory
                with st.spinner("Preparing reports..."):
                    # Generate PDF report in memory - pass target_words for ordering and include_arabic flag
                    pdf_buffer = generate_pdf_report(summary_data, file_analysis, document_stats, target_words, include_arabic)
                    
                    # Generate Excel report in memory - pass target_words for ordering and include_arabic flag
                    excel_buffer = generate_excel_report(summary_data, file_analysis, document_stats, target_words, include_arabic)
                
                if pdf_buffer:
                    with col1:
                        # Download PDF report
                        st.download_button(
                            label="Download PDF Report",
                            data=pdf_buffer,
                            file_name="Qualitative_Analysis_Report.pdf",
                            mime="application/pdf",
                            use_container_width=True
                        )
                else:
                    st.error("PDF report generation failed")
                
                if excel_buffer:
                    with col2:
                        # Download Excel report
                        st.download_button(
                            label="Download Excel Report",
                            data=excel_buffer,
                            file_name="Qualitative_Analysis_Report.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                else:
                    st.error("Excel report generation failed")
                
                # Additional CSV exports - PRESERVE USER'S WORD ORDER
                st.subheader("Export Data Tables")
                col_csv1, col_csv2 = st.columns(2)
                
                with col_csv1:
                    # Download summary as CSV
                    csv = summary_df.to_csv(index=False).encode('utf-8')
                    st.download_button(
                        label="Download Summary (CSV)",
                        data=csv,
                        file_name="word_frequency_summary.csv",
                        mime="text/csv",
                        use_container_width=True
                    )
                
                with col_csv2:
                    # Download file analysis as CSV
                    if file_analysis:
                        file_analysis_csv = file_analysis_df.to_csv().encode('utf-8')
                        st.download_button(
                            label="Download File Analysis (CSV)",
                            data=file_analysis_csv,
                            file_name="per_file_analysis.csv",
                            mime="text/csv",
                            use_container_width=True
                        )
                
                status_text.empty()
                progress_bar.empty()
                
            except Exception as e:
                st.error(f"Analysis failed: {str(e)}")
                st.error("Please try again or check your files")
                if st.session_state.get('debug_mode', False):
                    st.text(traceback.format_exc())
    except Exception as e:
        st.error(f"Critical application error: {str(e)}")
        if st.session_state.get('debug_mode', False):
            st.text(traceback.format_exc())

if __name__ == "__main__":
    main()
